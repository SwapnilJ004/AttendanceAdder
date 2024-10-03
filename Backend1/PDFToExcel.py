import convertapi
import sys
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

convertapi.api_credentials = 'secret_dNB0v09AfZ2yOwUe'
file_name = "./Btech_Attendance_2024"

convertapi.convert('xlsx', {
    'File': file_name + ".pdf"
}, from_format = 'pdf').save_files('./')

wb = load_workbook(file_name+".xlsx")

ws = wb.active

attendance_col = ws['F']
total_col = ws['E']

for i in attendance_col.length:
    print(i.value*100/j.value)